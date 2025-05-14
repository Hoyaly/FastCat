import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "User Info Test Cases"

# Define styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define headers
headers = [
    "Test ID",
    "Description",
    "Precondition",
    "Test Steps",
    "Expected Result",
    "Status",
    "Notes"
]

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Test cases data
test_cases = [
    {
        "Test ID": "TC_INFO_001",
        "Description": "Verify successful update of user information with valid data",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Update name: 'John Doe'\n3. Update phone: '0123456789'\n4. Update address: '123 Main Street'\n5. Click 'Save' button",
        "Expected Result": "- Information is successfully updated\n- Success message is displayed\n- Updated information is shown on the page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_002",
        "Description": "Verify update with empty name field",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Clear name field\n3. Update phone: '0123456789'\n4. Update address: '123 Main Street'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Name is required' is displayed\n- Information is not updated\n- Form remains in edit mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_003",
        "Description": "Verify update with empty phone field",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Update name: 'John Doe'\n3. Clear phone field\n4. Update address: '123 Main Street'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Phone number is required' is displayed\n- Information is not updated\n- Form remains in edit mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_004",
        "Description": "Verify update with invalid phone number format",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Update name: 'John Doe'\n3. Enter phone: 'abc123'\n4. Update address: '123 Main Street'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Invalid phone number format' is displayed\n- Information is not updated\n- Form remains in edit mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_005",
        "Description": "Verify update with empty address field",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Update name: 'John Doe'\n3. Update phone: '0123456789'\n4. Clear address field\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Address is required' is displayed\n- Information is not updated\n- Form remains in edit mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_006",
        "Description": "Verify cancel button functionality",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Make some changes to the information\n3. Click 'Cancel' button",
        "Expected Result": "- Edit mode is cancelled\n- Original information is restored\n- Form returns to view mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_007",
        "Description": "Verify update with very long name",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Enter a very long name (more than 50 characters)\n3. Update phone: '0123456789'\n4. Update address: '123 Main Street'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Name is too long' is displayed\n- Information is not updated\n- Form remains in edit mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_008",
        "Description": "Verify update with very long address",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Update name: 'John Doe'\n3. Update phone: '0123456789'\n4. Enter a very long address (more than 200 characters)\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Address is too long' is displayed\n- Information is not updated\n- Form remains in edit mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_009",
        "Description": "Verify update with special characters in name",
        "Precondition": "User is logged in and on the info.html page",
        "Test Steps": "1. Click 'Edit' button\n2. Enter name with special characters: 'John@Doe#123'\n3. Update phone: '0123456789'\n4. Update address: '123 Main Street'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Name contains invalid characters' is displayed\n- Information is not updated\n- Form remains in edit mode",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_INFO_010",
        "Description": "Verify information persistence after page refresh",
        "Precondition": "User is logged in and has updated their information",
        "Test Steps": "1. Update user information\n2. Refresh the page\n3. Check the displayed information",
        "Expected Result": "- Updated information is still displayed correctly\n- All fields show the last saved values",
        "Status": "",
        "Notes": ""
    }
]

# Write test cases
for row, test_case in enumerate(test_cases, 2):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = test_case[header]
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Adjust column widths
column_widths = {
    'A': 15,  # Test ID
    'B': 40,  # Description
    'C': 30,  # Precondition
    'D': 40,  # Test Steps
    'E': 40,  # Expected Result
    'F': 15,  # Status
    'G': 30   # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
wb.save('User_Info_Test_Cases.xlsx') 