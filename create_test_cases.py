import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Admin Login Test Cases"

# Define styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define headers
headers = [
    "Test ID",
    "Description",
    "Precondition",
    "Test Steps",
    "Expected Result",
    "Status",
    "Notes"
]

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Test cases data
test_cases = [
    {
        "Test ID": "TC_ADMIN_001",
        "Description": "Verify successful login with valid admin credentials",
        "Precondition": "User is on the login page (index.html)",
        "Test Steps": "1. Enter email: 'hoyaly123@gmail.com'\n2. Enter password: 'hoyaly'\n3. Click 'Đăng nhập' button",
        "Expected Result": "- User should be redirected to home.html\n- User should have admin privileges\n- User menu should show admin options (Quản lý, Thông tin cá nhân, Đăng xuất)",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADMIN_002",
        "Description": "Verify login attempt with invalid admin email",
        "Precondition": "User is on the login page (index.html)",
        "Test Steps": "1. Enter email: 'wrongadmin@gmail.com'\n2. Enter password: 'hoyaly'\n3. Click 'Đăng nhập' button",
        "Expected Result": "- Error message 'Email hoặc mật khẩu không đúng' should be displayed\n- User should remain on login page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADMIN_003",
        "Description": "Verify login attempt with invalid admin password",
        "Precondition": "User is on the login page (index.html)",
        "Test Steps": "1. Enter email: 'hoyaly123@gmail.com'\n2. Enter password: 'wrongpassword'\n3. Click 'Đăng nhập' button",
        "Expected Result": "- Error message 'Email hoặc mật khẩu không đúng' should be displayed\n- User should remain on login page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADMIN_004",
        "Description": "Verify admin login with 'Ghi nhớ đăng nhập' option",
        "Precondition": "User is on the login page (index.html)",
        "Test Steps": "1. Enter email: 'hoyaly123@gmail.com'\n2. Enter password: 'hoyaly'\n3. Check 'Ghi nhớ đăng nhập' checkbox\n4. Click 'Đăng nhập' button\n5. Close browser and reopen\n6. Navigate to index.html",
        "Expected Result": "- User should be automatically logged in\n- Email and password fields should be pre-filled\n- 'Ghi nhớ đăng nhập' checkbox should be checked",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADMIN_005",
        "Description": "Verify admin access to admin panel",
        "Precondition": "Admin is logged in",
        "Test Steps": "1. Click on user icon\n2. Click on 'Quản lý' option",
        "Expected Result": "- User should be redirected to admin.html\n- Admin panel should be accessible",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADMIN_006",
        "Description": "Verify non-admin user cannot access admin panel",
        "Precondition": "Regular user is logged in",
        "Test Steps": "1. Try to access admin.html directly",
        "Expected Result": "- Alert message 'Bạn không có quyền truy cập trang này' should be displayed\n- User should be redirected to home.html",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADMIN_007",
        "Description": "Verify admin logout functionality",
        "Precondition": "Admin is logged in",
        "Test Steps": "1. Click on user icon\n2. Click on 'Đăng xuất' option",
        "Expected Result": "- User should be logged out\n- User should be redirected to index.html\n- All admin privileges should be revoked",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADMIN_008",
        "Description": "Verify login attempt with invalid email format",
        "Precondition": "User is on the login page (index.html)",
        "Test Steps": "1. Enter email: 'hoyaly123@yahoo.com' (non-gmail email)\n2. Enter password: 'hoyaly'\n3. Click 'Đăng nhập' button",
        "Expected Result": "- Error message 'Email không đúng định dạng, vui lòng nhập lại email có đuôi @gmail.com' should be displayed\n- User should remain on login page",
        "Status": "",
        "Notes": ""
    }
]

# Write test cases
for row, test_case in enumerate(test_cases, 2):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = test_case[header]
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Adjust column widths
column_widths = {
    'A': 15,  # Test ID
    'B': 40,  # Description
    'C': 30,  # Precondition
    'D': 40,  # Test Steps
    'E': 40,  # Expected Result
    'F': 15,  # Status
    'G': 30   # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
wb.save('Admin_Login_Test_Cases.xlsx')

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Registration Test Cases"

# Define styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define headers
headers = [
    "Test ID",
    "Description",
    "Precondition",
    "Test Steps",
    "Expected Result",
    "Status",
    "Notes"
]

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Test cases data
test_cases = [
    {
        "Test ID": "TC_REG_001",
        "Description": "Verify successful registration with valid credentials",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Đăng ký' button",
        "Expected Result": "- Registration should be successful\n- User should be redirected to login page\n- Alert message 'Đăng ký thành công' should be displayed",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_002",
        "Description": "Verify registration with non-gmail email",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@yahoo.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Đăng ký' button",
        "Expected Result": "- Error message 'Email không đúng định dạng, vui lòng nhập lại email có đuôi @gmail.com' should be displayed\n- User should remain on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_003",
        "Description": "Verify registration with password less than 6 characters",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@gmail.com'\n2. Enter password: '12345'\n3. Enter confirm password: '12345'\n4. Click 'Đăng ký' button",
        "Expected Result": "- Error message 'Mật khẩu phải có ít nhất 6 ký tự' should be displayed\n- User should remain on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_004",
        "Description": "Verify registration with non-matching passwords",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password456'\n4. Click 'Đăng ký' button",
        "Expected Result": "- Error message 'Mật khẩu xác nhận không khớp' should be displayed\n- User should remain on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_005",
        "Description": "Verify registration with existing email",
        "Precondition": "User is on the registration page (register.html) and an account with email 'existing@gmail.com' already exists",
        "Test Steps": "1. Enter email: 'existing@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Đăng ký' button",
        "Expected Result": "- Error message 'Email này đã được đăng ký. Vui lòng sử dụng email khác.' should be displayed\n- User should remain on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_006",
        "Description": "Verify registration with empty fields",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Leave all fields empty\n2. Click 'Đăng ký' button",
        "Expected Result": "- Browser should show required field validation\n- User should remain on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_007",
        "Description": "Verify registration with admin email",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'hoyaly123@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Đăng ký' button",
        "Expected Result": "- Error message 'Không thể đăng ký tài khoản admin' should be displayed\n- User should remain on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_008",
        "Description": "Verify registration with special characters in email",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'test!user@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Đăng ký' button",
        "Expected Result": "- Error message 'Email không đúng định dạng, vui lòng nhập lại email có đuôi @gmail.com' should be displayed\n- User should remain on registration page",
        "Status": "",
        "Notes": ""
    }
]

# Write test cases
for row, test_case in enumerate(test_cases, 2):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = test_case[header]
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Adjust column widths
column_widths = {
    'A': 15,  # Test ID
    'B': 40,  # Description
    'C': 30,  # Precondition
    'D': 40,  # Test Steps
    'E': 40,  # Expected Result
    'F': 15,  # Status
    'G': 30   # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
wb.save('Registration_Test_Cases.xlsx') 