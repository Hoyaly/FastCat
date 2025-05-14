import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Shopping Cart Test Cases"

# Define styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define headers
headers = [
    "Test ID",
    "Description",
    "Precondition",
    "Test Steps",
    "Expected Result",
    "Status",
    "Notes"
]

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Test cases data
test_cases = [
    {
        "Test ID": "TC_CART_001",
        "Description": "Verify adding a product to empty cart",
        "Precondition": "User is logged in and cart is empty",
        "Test Steps": "1. Navigate to product page\n2. Click 'Add to Cart' button for 'Royal Canin Cat Food'\n3. Navigate to cart page",
        "Expected Result": "- Product is added to cart\n- Cart shows 1 item\n- Product details (name, price, quantity) are displayed correctly",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_002",
        "Description": "Verify adding multiple quantities of same product",
        "Precondition": "User is logged in and cart is empty",
        "Test Steps": "1. Navigate to product page\n2. Click 'Add to Cart' button for 'Royal Canin Cat Food' 3 times\n3. Navigate to cart page",
        "Expected Result": "- Product is added to cart\n- Cart shows 1 item with quantity 3\n- Total price is calculated correctly (price × 3)",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_003",
        "Description": "Verify adding different products to cart",
        "Precondition": "User is logged in and cart is empty",
        "Test Steps": "1. Navigate to product page\n2. Click 'Add to Cart' for 'Royal Canin Cat Food'\n3. Click 'Add to Cart' for 'Whiskas Cat Food'\n4. Navigate to cart page",
        "Expected Result": "- Both products are added to cart\n- Cart shows 2 items\n- Each product shows correct details\n- Total price is sum of both products",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_004",
        "Description": "Verify increasing product quantity in cart",
        "Precondition": "User is logged in and has 'Royal Canin Cat Food' in cart with quantity 1",
        "Test Steps": "1. Navigate to cart page\n2. Click '+' button next to 'Royal Canin Cat Food' quantity",
        "Expected Result": "- Product quantity increases to 2\n- Total price updates correctly\n- Cart total updates correctly",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_005",
        "Description": "Verify decreasing product quantity in cart",
        "Precondition": "User is logged in and has 'Royal Canin Cat Food' in cart with quantity 2",
        "Test Steps": "1. Navigate to cart page\n2. Click '-' button next to 'Royal Canin Cat Food' quantity",
        "Expected Result": "- Product quantity decreases to 1\n- Total price updates correctly\n- Cart total updates correctly",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_006",
        "Description": "Verify removing product when quantity reaches zero",
        "Precondition": "User is logged in and has 'Royal Canin Cat Food' in cart with quantity 1",
        "Test Steps": "1. Navigate to cart page\n2. Click '-' button next to 'Royal Canin Cat Food' quantity",
        "Expected Result": "- Product is removed from cart\n- Cart shows empty or remaining items\n- Cart total updates correctly",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_007",
        "Description": "Verify removing product using delete button",
        "Precondition": "User is logged in and has multiple products in cart",
        "Test Steps": "1. Navigate to cart page\n2. Click 'Delete' button next to 'Royal Canin Cat Food'",
        "Expected Result": "- Product is removed from cart\n- Remaining products are still displayed\n- Cart total updates correctly",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_008",
        "Description": "Verify cart persistence after page refresh",
        "Precondition": "User is logged in and has products in cart",
        "Test Steps": "1. Navigate to cart page\n2. Note the products and quantities\n3. Refresh the page",
        "Expected Result": "- All products remain in cart\n- Quantities remain unchanged\n- Total price remains correct",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_009",
        "Description": "Verify cart persistence after logout and login",
        "Precondition": "User is logged in and has products in cart",
        "Test Steps": "1. Navigate to cart page\n2. Note the products and quantities\n3. Logout\n4. Login with same account\n5. Navigate to cart page",
        "Expected Result": "- All products remain in cart\n- Quantities remain unchanged\n- Total price remains correct",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_CART_010",
        "Description": "Verify cart total calculation with multiple items",
        "Precondition": "User is logged in and has multiple products with different quantities in cart",
        "Test Steps": "1. Navigate to cart page\n2. Add multiple products with different quantities\n3. Verify the total calculation",
        "Expected Result": "- Cart total is sum of (price × quantity) for each product\n- Total updates correctly when quantities change\n- Total updates correctly when products are removed",
        "Status": "",
        "Notes": ""
    }
]

# Write test cases
for row, test_case in enumerate(test_cases, 2):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = test_case[header]
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Adjust column widths
column_widths = {
    'A': 15,  # Test ID
    'B': 40,  # Description
    'C': 30,  # Precondition
    'D': 40,  # Test Steps
    'E': 40,  # Expected Result
    'F': 15,  # Status
    'G': 30   # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
wb.save('Shopping_Cart_Test_Cases.xlsx') 