import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Registration Test Cases"

# Define styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define headers
headers = [
    "Test ID",
    "Description",
    "Precondition",
    "Test Steps",
    "Expected Result",
    "Status",
    "Notes"
]

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Test cases data
test_cases = [
    {
        "Test ID": "TC_REG_001",
        "Description": "Verify successful registration with valid credentials",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Register' button",
        "Expected Result": "- Registration successful\n- Redirected to login page\n- Display success message 'Registration successful'",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_002",
        "Description": "Verify registration with non-gmail email",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@yahoo.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'Invalid email format, please enter a Gmail address'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_003",
        "Description": "Verify registration with password less than 6 characters",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@gmail.com'\n2. Enter password: '12345'\n3. Enter confirm password: '12345'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'Password must be at least 6 characters'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_004",
        "Description": "Verify registration with non-matching passwords",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password456'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'Passwords do not match'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_005",
        "Description": "Verify registration with existing email",
        "Precondition": "User is on the registration page (register.html) and an account with email 'existing@gmail.com' already exists",
        "Test Steps": "1. Enter email: 'existing@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'This email is already registered. Please use a different email.'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_006",
        "Description": "Verify registration with empty fields",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Leave all fields empty\n2. Click 'Register' button",
        "Expected Result": "- Browser should show required field validation\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_007",
        "Description": "Verify registration with admin email",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'hoyaly123@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'Cannot register admin account'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_008",
        "Description": "Verify registration with email containing special characters",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'test!user@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'Invalid email format, please enter a Gmail address'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_009",
        "Description": "Verify registration with email containing spaces",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'test user@gmail.com'\n2. Enter password: 'password123'\n3. Enter confirm password: 'password123'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'Invalid email format, please enter a Gmail address'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_REG_010",
        "Description": "Verify registration with password containing spaces",
        "Precondition": "User is on the registration page (register.html)",
        "Test Steps": "1. Enter email: 'testuser@gmail.com'\n2. Enter password: 'pass word123'\n3. Enter confirm password: 'pass word123'\n4. Click 'Register' button",
        "Expected Result": "- Display error message 'Password cannot contain spaces'\n- Stay on registration page",
        "Status": "",
        "Notes": ""
    }
]

# Write test cases
for row, test_case in enumerate(test_cases, 2):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = test_case[header]
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Adjust column widths
column_widths = {
    'A': 15,  # Test ID
    'B': 40,  # Description
    'C': 30,  # Precondition
    'D': 40,  # Test Steps
    'E': 40,  # Expected Result
    'F': 15,  # Status
    'G': 30   # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
wb.save('Registration_Test_Cases.xlsx') 