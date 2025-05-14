import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Admin Product Test Cases"

# Define styles
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define headers
headers = [
    "Test ID",
    "Description",
    "Precondition",
    "Test Steps",
    "Expected Result",
    "Status",
    "Notes"
]

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Test cases data
test_cases = [
    {
        "Test ID": "TC_ADM_PROD_001",
        "Description": "Verify successful product addition with valid data",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Enter price: '250000'\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Save' button",
        "Expected Result": "- Product is successfully added\n- New product appears in the product list\n- Success message is displayed",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_002",
        "Description": "Verify product addition with empty name field",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Leave product name empty\n3. Enter price: '250000'\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Product name is required' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_003",
        "Description": "Verify product addition with empty price field",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Leave price empty\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Price is required' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_004",
        "Description": "Verify product addition with invalid price format",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Enter price: 'abc'\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Price must be a valid number' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_005",
        "Description": "Verify product addition with negative price",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Enter price: '-250000'\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Price must be greater than 0' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_006",
        "Description": "Verify product addition with empty image URL",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Enter price: '250000'\n4. Leave image URL empty\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Image URL is required' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_007",
        "Description": "Verify product addition with invalid image URL format",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Enter price: '250000'\n4. Enter image URL: 'invalid-url'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Invalid image URL format' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_008",
        "Description": "Verify product addition with duplicate product name",
        "Precondition": "Admin is logged in and on the admin product management page, and a product named 'Royal Canin Cat Food' already exists",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Enter price: '250000'\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Product with this name already exists' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_009",
        "Description": "Verify product addition with very long product name",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'A very long product name that exceeds the maximum allowed length for product names in the system'\n3. Enter price: '250000'\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Save' button",
        "Expected Result": "- Error message 'Product name is too long' is displayed\n- Product is not added\n- Form remains open",
        "Status": "",
        "Notes": ""
    },
    {
        "Test ID": "TC_ADM_PROD_010",
        "Description": "Verify cancel button functionality",
        "Precondition": "Admin is logged in and on the admin product management page",
        "Test Steps": "1. Click 'Add Product' button\n2. Enter product name: 'Royal Canin Cat Food'\n3. Enter price: '250000'\n4. Enter image URL: 'https://example.com/catfood.jpg'\n5. Click 'Cancel' button",
        "Expected Result": "- Add product form is closed\n- No product is added\n- Admin returns to product list view",
        "Status": "",
        "Notes": ""
    }
]

# Write test cases
for row, test_case in enumerate(test_cases, 2):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = test_case[header]
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Adjust column widths
column_widths = {
    'A': 15,  # Test ID
    'B': 40,  # Description
    'C': 30,  # Precondition
    'D': 40,  # Test Steps
    'E': 40,  # Expected Result
    'F': 15,  # Status
    'G': 30   # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
wb.save('Admin_Product_Test_Cases.xlsx') 